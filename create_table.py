"""
Convert extracted JSON data to multi-tab Excel format for analysis
Creates one tab per extracted field block
"""

import difflib
import json
import os
import re
from pathlib import Path


def load_extracted_data(filepath):
    """Load the extracted JSON data"""
    print(f"[*] Loading extracted data from: {filepath}")
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"[OK] Data loaded successfully")
        return data
    except Exception as e:
        print(f"[ERROR] Failed to load data: {e}")
        raise


def create_metadata_sheet(workbook, metadata):
    """Create metadata tab with Carrier and Validity info"""
    print("[*] Creating Metadata tab...")
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    ws = workbook.create_sheet("Metadata", 0)
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Data
    data = [
        ["Field", "Value"],
        ["Client", metadata.get('client', '')],
        ["Carrier", metadata.get('carrier', '').replace('\n', ' ')],
        ["Validity Date", metadata.get('validity_date', '')],
        ["Extraction Date", metadata.get('extraction_date', '')],
        ["Extraction Source", metadata.get('extraction_source', '')]
    ]
    
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    
    print(f"[OK] Metadata tab created")


def pivot_main_costs(main_costs, metadata):
    """Pivot MainCosts data - zones as rows, weights as columns (legacy flat view)"""
    rows = []
    client = metadata.get('client', '')
    carrier = metadata.get('carrier', '').replace('\n', ' ')
    validity_date = metadata.get('validity_date', '')
    
    for section_idx, rate_card in enumerate(main_costs, 1):
        service_type = rate_card.get('service_type') or ''
        cost_category = rate_card.get('cost_category', '')
        weight_unit = rate_card.get('weight_unit', 'KG')
        zone_headers = rate_card.get('zone_headers', {})
        pricing = rate_card.get('pricing', [])
        
        # Build a matrix: zone -> weight -> price
        zone_price_matrix = {}
        weights_set = set()
        
        for price_entry in pricing:
            weight = price_entry.get('weight', '')
            weights_set.add(weight)
            zone_prices = price_entry.get('zone_prices', {})
            
            for zone_key, price in zone_prices.items():
                zone_name = zone_headers.get(zone_key, zone_key)
                if zone_name not in zone_price_matrix:
                    zone_price_matrix[zone_name] = {}
                zone_price_matrix[zone_name][weight] = price
        
        # Sort weights numerically
        try:
            weights_sorted = sorted(weights_set, key=lambda x: float(x))
        except:
            weights_sorted = sorted(weights_set)
        
        # Create rows for each zone
        for zone_name, weight_prices in zone_price_matrix.items():
            row = {
                'Client': client,
                'Carrier': carrier,
                'Validity Date': validity_date,
                'Section': section_idx,
                'Service Type': service_type,
                'Cost Category': cost_category,
                'Weight Unit': weight_unit,
                'Zone': zone_name
            }
            
            # Add prices for each weight as columns
            for weight in weights_sorted:
                col_name = f"{weight} {weight_unit}"
                row[col_name] = weight_prices.get(weight, '')
            
            rows.append(row)
    
    return rows


def _zone_has_letters(zone_name):
    """True if the zone identifier contains a letter (e.g. 'Zone A' yes, 'Zone 1' no)."""
    s = (zone_name or '').strip()
    if not s:
        return False
    # Strip leading "Zone " so we look at the identifier part only
    if s.upper().startswith('ZONE '):
        suffix = s[5:].strip()
    else:
        suffix = s
    return any(c.isalpha() for c in suffix)


def parse_zoning_matrix(zoning_matrix):
    """
    Parse ZoningMatrix list into a lookup for origin/destination by matrix and zone letter.
    Structure: rows with MatrixName start a new matrix (header row with dest zone numbers);
    following rows with OriginZone are data rows (cell value = zone letter).
    Returns: dict[matrix_name, str] -> list of (origin_zone, dest_zone) tuples.
    Key is (matrix_name_normalized, zone_letter) where zone_letter is e.g. 'A', 'E'.
    """
    import re
    result = {}  # (matrix_name, zone_letter) -> [(origin_zone, dest_zone), ...]
    dest_cols = None  # list of 'DestinationZone1', 'DestinationZone2', ... in order
    header_dest_nums = None  # list of zone numbers for each column (from header row)
    current_matrix_name = None

    for row in zoning_matrix or []:
        matrix_name = (row.get('MatrixName') or '').strip()
        origin_zone = (row.get('OriginZone') or '').strip()

        if matrix_name:
            # Start of a new matrix: header row
            current_matrix_name = matrix_name
            dest_keys = sorted(
                [k for k in row if re.match(r'^DestinationZone\d+$', k)],
                key=lambda k: int(re.search(r'\d+', k).group())
            )
            dest_cols = dest_keys
            header_dest_nums = [str(row.get(k, '')).strip() for k in dest_cols]
            continue

        if current_matrix_name and origin_zone and dest_cols:
            # Data row: OriginZone + DestinationZone1..N with letters
            for col_idx, dest_key in enumerate(dest_cols):
                if col_idx >= len(header_dest_nums):
                    continue
                dest_zone_num = header_dest_nums[col_idx]
                if not dest_zone_num:
                    continue
                cell_letter = (row.get(dest_key) or '').strip()
                if not cell_letter:
                    continue
                key = (current_matrix_name, cell_letter.upper())
                if key not in result:
                    result[key] = []
                result[key].append((origin_zone, dest_zone_num))

    return result


def _matrix_zone_to_letter(matrix_zone):
    """Extract zone letter from 'Zone E' -> 'E'."""
    s = (matrix_zone or '').strip()
    if not s:
        return ''
    if s.upper().startswith('ZONE '):
        return s[5:].strip().upper()
    return s.upper()


def _main_words(text):
    """Return set of significant words (uppercase) from text, excluding ZONE/MATRIX."""
    if not text:
        return set()
    words = set((text or '').upper().split())
    words.discard('ZONE')
    words.discard('MATRIX')
    return words


def _find_matrix_for_service(zoning_lookup, service):
    """
    Return a matrix name from zoning_lookup that matches service.
    Matches when: (1) service is substring of matrix name or vice versa, or
    (2) all main words of the matrix name (excluding ZONE, MATRIX) are present in the service.
    E.g. Service 'DHL EXPRESS WORLDWIDE THIRD COUNTRY' matches matrix
    'DHL EXPRESS THIRD COUNTRY ZONE MATRIX' because DHL, EXPRESS, THIRD, COUNTRY are in both.
    """
    service = (service or '').strip()
    if not service:
        return None
    service_words = _main_words(service)
    matrix_names = {mn for (mn, _) in zoning_lookup}

    for mn in matrix_names:
        if service in mn or mn in service:
            return mn
    for mn in matrix_names:
        normalized = mn.replace(' ZONE MATRIX', '').strip()
        if service in normalized or normalized in service:
            return mn
    # Main-words match: all significant words from matrix name appear in service
    for mn in matrix_names:
        matrix_words = _main_words(mn.replace(' ZONE MATRIX', ''))
        if matrix_words and matrix_words <= service_words:
            return mn
    return None


def expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix):
    """
    For rows where Matrix zone is not empty, look up (Origin, Destination) pairs from
    ZoningMatrix and duplicate the lane for each pair. Re-number Lane # at the end.
    """
    if not matrix_rows:
        return matrix_rows
    zoning_lookup = parse_zoning_matrix(zoning_matrix)
    if not zoning_lookup:
        return matrix_rows

    expanded = []
    for row in matrix_rows:
        matrix_zone = (row.get('Matrix zone') or '').strip()
        service = (row.get('Service') or '').strip()

        if not matrix_zone:
            expanded.append(row)
            continue

        zone_letter = _matrix_zone_to_letter(matrix_zone)
        if not zone_letter:
            expanded.append(row)
            continue

        matrix_name = _find_matrix_for_service(zoning_lookup, service)
        if not matrix_name:
            expanded.append(row)
            continue

        key = (matrix_name, zone_letter)
        pairs = zoning_lookup.get(key, [])
        if not pairs:
            expanded.append(row)
            continue

        for origin_zone, dest_zone in pairs:
            new_row = row.copy()
            new_row['Origin'] = f"Zone {origin_zone}" if origin_zone else ''
            new_row['Destination'] = f"Zone {dest_zone}" if dest_zone else ''
            expanded.append(new_row)

    # Re-number Lane #
    for lane, row in enumerate(expanded, 1):
        row['Lane #'] = lane

    return expanded


def global_country(metadata):
    """Last word from Carrier name, e.g. 'DHL DHL Express France' -> 'France'."""
    carrier = (metadata.get('carrier') or '').replace('\n', ' ').strip()
    parts = carrier.split()
    return parts[-1] if parts else ''


def _zone_sort_key(zone_name):
    """Sort key for zone names: 'Zone 1' < 'Zone 2' < ... < 'Zone 10' < 'Zone A'."""
    s = (zone_name or '').strip()
    if not s:
        return (1, 0)
    if s.upper().startswith('ZONE '):
        suffix = s[5:].strip()
    else:
        suffix = s
    try:
        return (0, float(suffix))
    except (ValueError, TypeError):
        return (1, 0 if _zone_has_letters(zone_name) else 1)


def build_matrix_main_costs(main_costs, metadata):
    """
    Build the Matrix view: one row per lane (service + zone). All cost categories
    (Envelope, Documents, etc.) for the same Service + Destination/Origin are
    merged into a single row with multiple cost columns.
    Returns (rows, category_specs). category_specs = [(cost_cat_name, weight_unit, [weights]), ...].
    """
    # Unique cost categories in order of first appearance, with weight_unit and sorted weights
    category_specs = []  # (cost_category_name, weight_unit, [weight1, weight2, ...])
    seen_categories = {}

    for rate_card in main_costs:
        cost_category = rate_card.get('cost_category') or ''
        weight_unit = rate_card.get('weight_unit') or 'KG'
        pricing = rate_card.get('pricing', [])
        weights_set = set()
        for pe in pricing:
            w = pe.get('weight', '')
            if w:
                weights_set.add(w)
        try:
            weights_sorted = sorted(weights_set, key=lambda x: float(x))
        except (ValueError, TypeError):
            weights_sorted = sorted(weights_set)
        if cost_category not in seen_categories:
            seen_categories[cost_category] = (weight_unit, weights_sorted)
            category_specs.append((cost_category, weight_unit, weights_sorted))
        else:
            # Merge weights if same category appears again with different weights
            existing_unit, existing_weights = seen_categories[cost_category]
            merged = set(existing_weights) | set(weights_sorted)
            try:
                merged_sorted = sorted(merged, key=lambda x: float(x))
            except (ValueError, TypeError):
                merged_sorted = sorted(merged)
            seen_categories[cost_category] = (existing_unit, merged_sorted)
            # Update in category_specs
            for i, spec in enumerate(category_specs):
                if spec[0] == cost_category:
                    category_specs[i] = (cost_category, existing_unit, merged_sorted)
                    break

    # One row per (service_type, zone_name); merge all cost categories into that row
    # Key: (service_type, zone_name) -> row dict (without Lane # yet)
    lane_rows = {}

    for rate_card in main_costs:
        service_type = (rate_card.get('service_type') or '').strip()
        cost_category = rate_card.get('cost_category') or ''
        zone_headers = rate_card.get('zone_headers', {})
        pricing = rate_card.get('pricing', [])
        service_lower = service_type.lower()
        is_import = 'import' in service_lower
        is_export = 'export' in service_lower

        zone_price_matrix = {}
        for price_entry in pricing:
            weight = price_entry.get('weight', '')
            zone_prices = price_entry.get('zone_prices', {})
            for zone_key, price in zone_prices.items():
                zone_name = zone_headers.get(zone_key, zone_key)
                if zone_name not in zone_price_matrix:
                    zone_price_matrix[zone_name] = {}
                zone_price_matrix[zone_name][weight] = price

        for zone_name, weight_prices in zone_price_matrix.items():
            key = (service_type, zone_name)
            if key not in lane_rows:
                origin = zone_name if is_import else ''
                destination = zone_name if is_export else ''
                matrix_zone = zone_name if _zone_has_letters(zone_name) else ''
                lane_rows[key] = {
                    'Origin': origin,
                    'Destination': destination,
                    'Service': service_type,
                    'Matrix zone': matrix_zone,
                }
            row = lane_rows[key]
            for weight, price in weight_prices.items():
                row[(cost_category, weight)] = price

    # Carrier last word: used for DOMESTIC and for filling empty Origin/Destination when Matrix zone is empty
    carrier_last = global_country(metadata)

    # Sort by service_type then zone, assign Lane # and build list
    sorted_keys = sorted(lane_rows.keys(), key=lambda k: (k[0], _zone_sort_key(k[1])))
    rows = []
    for lane, key in enumerate(sorted_keys, 1):
        row = lane_rows[key].copy()
        row['Lane #'] = lane

        # Origin/Destination from carrier when applicable
        service = (row.get('Service') or '').strip()
        matrix_zone = (row.get('Matrix zone') or '').strip()
        if service == 'DHL EXPRESS DOMESTIC':
            if carrier_last:
                row['Origin'] = carrier_last
                row['Destination'] = carrier_last
        elif not matrix_zone:
            # Matrix zone empty: fill empty Origin or Destination with carrier last word
            if carrier_last:
                if not (row.get('Origin') or '').strip():
                    row['Origin'] = carrier_last
                if not (row.get('Destination') or '').strip():
                    row['Destination'] = carrier_last

        rows.append(row)

    return rows, category_specs


def _transform_rate_name_to_short(rate_name):
    """
    Transform a rate card name into the short prefix for CountryZoning.
    Rules: remove DHL EXPRESS; INTERNATIONAL->WW; IMPORT->IMP; EXPORT->EXP;
    THIRD COUNTRY->3RD COUNTRY; Zoning->Zone. Domestic and Economy are kept.
    Collect tokens in order to form e.g. WW_EXP_IMP_ZONE_ (suffixed with zone).
    """
    if not rate_name or not isinstance(rate_name, str):
        return ''
    s = rate_name.upper().strip()
    # Apply replacements (order matters for multi-word)
    s = s.replace('DHL EXPRESS', ' ')
    s = s.replace('THIRD COUNTRY', ' 3RD_COUNTRY ')
    s = s.replace('INTERNATIONAL', ' WW ')
    s = s.replace('IMPORT', ' IMP ')
    s = s.replace('EXPORT', ' EXP ')
    s = s.replace('ZONING', ' ZONE ')
    # Collect tokens in fixed order: WW, 3RD_COUNTRY, DOMESTIC, ECONOMY, EXP, IMP, ZONE
    tokens = []
    for token in ('WW', '3RD_COUNTRY', 'DOMESTIC', 'ECONOMY', 'EXP', 'IMP', 'ZONE'):
        if token in s and token not in tokens:
            tokens.append(token)
    return '_'.join(tokens) if tokens else ''


def _fill_country_zoning_rate_names(rows):
    """
    Fill empty RateName cells in CountryZoning rows. Use last non-empty RateName,
    transform it to short form (e.g. WW_EXP_IMP_ZONE_), then append _<Zone>.
    """
    last_rate_name = ''
    for row in rows:
        rate_name = row.get('RateName') or ''
        zone = row.get('Zone') or ''
        if rate_name:
            last_rate_name = rate_name
        if not rate_name and last_rate_name and zone:
            prefix = _transform_rate_name_to_short(last_rate_name)
            if prefix:
                row['RateName'] = f"{prefix}_{zone}"


def _load_country_codes(codes_path=None):
    """
    Load country name -> code from dhl_country_codes.txt (format: Country\\tCode).
    If code contains a comma, the first code is used. Returns dict.
    """
    if codes_path is None:
        base = Path(__file__).resolve().parent
        codes_path = base / "input" / "dhl_country_codes.txt"
        if not codes_path.exists():
            codes_path = base / "addition" / "dhl_country_codes.txt"
    codes_path = Path(codes_path)
    if not codes_path.exists():
        return {}
    name_to_code = {}
    for line in codes_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or "\t" not in line:
            continue
        name, code = line.split("\t", 1)
        name = name.strip()
        code = code.strip()
        if "," in code:
            code = code.split(",")[0].strip()
        if name:
            name_to_code[name] = code
    return name_to_code


def _country_to_code(country, name_to_code):
    """
    Return code for country name; if not found, return empty string.
    Tries exact match, then normalizations for common differences:
    - Republic Of <-> Rep. Of
    - And <-> & (e.g. Bosnia And Herzegovina <-> Bosnia & Herzegovina)
    - Strip ", Peoples Republic" / ", People's Republic" (e.g. China, Peoples Republic -> China)
    """
    if not country:
        return ''
    s = str(country).strip()
    if not s:
        return ''
    # Exact and uppercase
    code = name_to_code.get(s)
    if code is not None:
        return code
    code = name_to_code.get(s.upper())
    if code is not None:
        return code
    # Build variants to try (order: base normalizations, then And/&, then strip suffixes)
    variants = []
    n = s.replace("Republic Of", "Rep. Of").replace("Republic of", "Rep. Of")
    n = n.replace(", Republic", ", Rep.").replace(" Republic", " Rep.")
    variants.append(n)
    variants.append(n.replace(" And ", " & "))
    variants.append(n.replace(" & ", " And "))
    # Strip ", Peoples Republic" / ", People's Republic" etc. (base name = e.g. China)
    for suffix in (", Peoples Republic", ", People's Republic", ", Peoples Rep.", ", People's Rep.",
                   " Peoples Republic", " People's Republic"):
        if n.endswith(suffix) or suffix in n:
            base = n.replace(suffix, "").strip().strip(",").strip()
            if base:
                variants.append(base)
    for v in variants:
        if not v:
            continue
        code = name_to_code.get(v)
        if code is not None:
            return code
        code = name_to_code.get(v.upper())
        if code is not None:
            return code
    return ''


def _fill_country_zoning_country_codes(rows, name_to_code):
    """Add Country Code column to CountryZoning rows from Country column."""
    for row in rows:
        country = row.get('Country') or ''
        row['Country Code'] = _country_to_code(country, name_to_code)


def flatten_array_data(array_data, metadata, field_name):
    """Flatten array data to table format"""
    rows = []
    client = metadata.get('client', '')
    carrier = metadata.get('carrier', '').replace('\n', ' ')
    validity_date = metadata.get('validity_date', '')
    
    for item in array_data:
        row = {
            'Client': client,
            'Carrier': carrier,
            'Validity Date': validity_date
        }
        row.update(item)
        rows.append(row)
    
    if field_name == 'CountryZoning':
        _fill_country_zoning_rate_names(rows)
        name_to_code = _load_country_codes()
        _fill_country_zoning_country_codes(rows, name_to_code)
    
    return rows


def _is_added_rates_header_row(item):
    """True if this item is a header row (zone names in Zone1..ZoneN), not a data row."""
    weight_from = item.get('WeightFrom', '')
    zone1_val = item.get('Zone1', '')
    if weight_from == 'From' or (str(zone1_val).strip().startswith('Zone')):
        return True
    return False


def pivot_added_rates(added_rates, metadata):
    """
    Convert AddedRates to table: rows = weight ranges (and header rows), columns = zones.
    Every JSON item becomes one row. Page Stopper and Table Name are filled only on rows
    where they exist in the JSON (the header row of each block); data rows leave them empty.
    """
    rows = []
    client = metadata.get('client', '')
    carrier = metadata.get('carrier', '').replace('\n', ' ')
    validity_date = metadata.get('validity_date', '')
    zone_column_names = []  # ordered list of (Zone1, "Zone 1"), (Zone2, "Zone 2"), ...

    for item in added_rates:
        is_header = _is_added_rates_header_row(item)
        if is_header:
            zone_column_names = []
            zone_keys = [k for k in item.keys() if k.startswith('Zone')]
            def zone_sort_key(k):
                suffix = k[4:]  # after 'Zone'
                try:
                    return int(suffix)
                except ValueError:
                    return 0
            for k in sorted(zone_keys, key=zone_sort_key):
                zone_column_names.append((k, str(item.get(k, k)).strip() or k))

        weight_from = item.get('WeightFrom', '')
        weight_to = item.get('WeightTo', '')
        row = {
            'Client': client,
            'Carrier': carrier,
            'Validity Date': validity_date,
            'Page Stopper': item.get('PageStopper', '') if is_header else '',
            'Table Name': item.get('TableName', '') if is_header else '',
            'Weight From': weight_from,
            'Weight To': weight_to,
        }
        for zone_key, zone_label in zone_column_names:
            row[zone_label] = item.get(zone_key, '')
        rows.append(row)
    return rows


def write_matrix_sheet(workbook, sheet_name, matrix_rows, category_specs, metadata):
    """Write the MainCosts matrix view: Row 1 = fixed + Cost Category; Row 2 = 'Weight measure - KG' only; Row 3 = weight brackets (0.3, 0.5, ...); then data rows."""
    if not matrix_rows:
        print(f"[WARN] No matrix data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} (Matrix) tab with {len(matrix_rows)} lanes...")
    ws = workbook.create_sheet(sheet_name)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fixed_cols = ['Lane #', 'Origin', 'Destination', 'Service', 'Matrix zone']
    num_fixed = len(fixed_cols)
    col = 1

    # Row 1: fixed headers + Cost Category merged groups
    for c, name in enumerate(fixed_cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    col = num_fixed + 1

    category_start_cols = []  # (start_col, end_col, cost_cat_name, weight_unit, weights)
    for cost_cat_name, weight_unit, weights in category_specs:
        start_col = col
        # Row 2: "Weight measure - KG" in first cell only; rest of category columns empty
        weight_measure_label = f"Weight measure - {weight_unit}" if weight_unit else "Weight measure"
        ws.cell(row=2, column=col, value=weight_measure_label)
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=2, column=col).font = header_font
        ws.cell(row=2, column=col).alignment = header_alignment
        col += 1
        for _ in weights:
            ws.cell(row=2, column=col, value='')
            ws.cell(row=2, column=col).fill = header_fill
            col += 1
        # Row 3: empty first cell, then weight brackets (0.3, 0.5, ...)
        ws.cell(row=3, column=start_col, value='')
        ws.cell(row=3, column=start_col).fill = header_fill
        col = start_col + 1
        for w in weights:
            ws.cell(row=3, column=col, value=w)
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font
            ws.cell(row=3, column=col).alignment = header_alignment
            col += 1
        end_col = col - 1
        category_start_cols.append((start_col, end_col, cost_cat_name, weight_unit, weights))
        # Row 1: merge and set Cost Category name
        if start_col <= end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=cost_cat_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    total_cols = col - 1

    # Row 2 & 3: fixed cells (under Lane #, Origin, etc.) empty with header fill
    for r in (2, 3):
        for c in range(1, num_fixed + 1):
            ws.cell(row=r, column=c, value='')
            ws.cell(row=r, column=c).fill = header_fill

    # Data rows start at row 4
    for row_idx, row_data in enumerate(matrix_rows, 4):
        col = 1
        for fc in fixed_cols:
            val = row_data.get(fc, '')
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fc == 'Lane #':
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            col += 1
        for start_col, end_col, cost_cat_name, weight_unit, weights in category_start_cols:
            # First column of group: leave empty in data
            cell = ws.cell(row=row_idx, column=start_col, value='')
            cell.alignment = Alignment(horizontal="center")
            col = start_col + 1
            for w in weights:
                key = (cost_cat_name, w)
                val = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                col += 1

    # Column widths
    last_data_row = len(matrix_rows) + 3
    for c in range(1, total_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10
        for r in range(1, min(last_data_row + 1, 54)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}{last_data_row}"
    print(f"[OK] {sheet_name} (Matrix) tab created with {total_cols} columns")


def write_sheet(workbook, sheet_name, rows, metadata):
    """Write data to a worksheet"""
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return
    
    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    ws = workbook.create_sheet(sheet_name)
    
    # Get all unique columns
    all_columns = set()
    for row in rows:
        all_columns.update(row.keys())
    
    # Define column order
    priority_cols = ['Client', 'Carrier', 'Validity Date', 'Section', 'Service Type', 
                     'Cost Category', 'Weight Unit', 'Zone', 'Page Stopper', 'Table Name',
                     'Weight From', 'Weight To',
                     'RateName', 'Country', 'Country Code', 'WeightFrom', 'WeightTo']
    
    columns = []
    for col in priority_cols:
        if col in all_columns:
            columns.append(col)
            all_columns.discard(col)
    
    # Separate weight columns from other remaining columns
    weight_cols = []
    other_cols = []
    
    for col in all_columns:
        if 'KG' in col or col.startswith('<=') or '-' in col:
            weight_cols.append(col)
        else:
            other_cols.append(col)
    
    # Sort weight columns numerically
    try:
        weight_cols_sorted = sorted(weight_cols, key=lambda x: float(x.split()[0]))
    except Exception:
        weight_cols_sorted = sorted(weight_cols)

    # Sort other_cols: Zone 1, Zone 2, ... Zone 16 before alphabetical
    def _other_col_sort_key(c):
        m = re.match(r'^Zone\s+(\d+)$', c, re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))
        return (1, c)
    columns.extend(weight_cols_sorted)
    columns.extend(sorted(other_cols, key=_other_col_sort_key))
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Center align certain columns
            if column in ['Weight', 'Weight Unit', 'Section', 'Zone', 'Currency', 'Rate'] or 'KG' in column:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        
        # Calculate max width
        max_length = len(str(column))
        for row_idx in range(2, min(len(rows) + 2, 52)):  # Check first 50 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (with limits)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add filter
    ws.auto_filter.ref = ws.dimensions
    
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


ACCESSORIAL_COSTS_COLUMNS = [
    'Original Cost Name',
    'Cost Type',
    'Cost Price',
    'Currency',
    'Rate by',
    'Apply Over',
    'Apply if',
    'Additional info(Cost Code)',
    'Valid From',
    'Valid To',
    'Carrier',
]


def _load_accessorial_cost_type_names(ref_path):
    """
    Load list of possible cost type names from addition/Accessorial Costs file.
    Supports .xlsx (first sheet) and .csv. Reads column 'Name'. Returns list of unique non-empty strings.
    """
    ref_path = Path(ref_path)
    if not ref_path.exists():
        return []
    names = []
    try:
        if ref_path.suffix.lower() in ('.xlsx', '.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
            ws = wb.active
            header = None
            name_col = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c).strip() if c is not None else '' for c in row]
                    for i, h in enumerate(header):
                        if h == 'Name':
                            name_col = i
                            break
                    if name_col is None:
                        break
                    continue
                if name_col is not None and name_col < len(row):
                    val = row[name_col]
                    if val is not None and str(val).strip():
                        names.append(str(val).strip())
            wb.close()
        elif ref_path.suffix.lower() == '.csv':
            import csv
            with open(ref_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header:
                    try:
                        name_col = header.index('Name')
                    except ValueError:
                        name_col = None
                    if name_col is not None:
                        for row in reader:
                            if name_col < len(row) and row[name_col].strip():
                                names.append(row[name_col].strip())
        else:
            return []
    except Exception:
        return []
    return list(dict.fromkeys(names))


def _token_set(text):
    """Split text into tokens (words and punctuation-separated parts like 9:00) for matching."""
    import re
    s = (text or '').lower().strip()
    tokens = set(re.findall(r'[a-z0-9]+(?::[a-z0-9]+)?|[a-z]+', s))
    return tokens


def _best_match_cost_type(original_name, name_list, cutoff=0.4):
    """
    Return the best matching name from name_list for original_name, or '' if none.
    Uses both character similarity (difflib) and token overlap so that e.g. "Premium 9:00:"
    matches "9:00 Service Fee" (shared token "9:00") better than "Permit Fee".
    Score = char_ratio + token_bonus where token_bonus rewards shared significant tokens.
    """
    if not original_name or not name_list:
        return ''
    original = str(original_name).strip()
    if not original:
        return ''
    orig_tokens = _token_set(original)
    best_score = -1.0
    best_name = ''
    for name in name_list:
        name_str = str(name).strip()
        if not name_str:
            continue
        char_ratio = difflib.SequenceMatcher(None, original.lower(), name_str.lower()).ratio()
        name_tokens = _token_set(name_str)
        shared = orig_tokens & name_tokens
        # Bonus when significant tokens match (e.g. "9:00", "premium", "fee"); ignore tiny tokens
        meaningful_orig = {t for t in orig_tokens if len(t) >= 2 or ':' in t}
        token_bonus = (len(shared & meaningful_orig) / len(meaningful_orig)) * 0.4 if meaningful_orig else 0.0
        score = char_ratio + token_bonus
        if score > best_score:
            best_score = score
            best_name = name_str
    return best_name if best_score >= 0.3 else ''


def build_accessorial_costs_rows(additional_costs_1, additional_costs_2, metadata, cost_type_ref_path=None):
    """
    Build rows for the Accessorial Costs tab from AdditionalCostsPart1 and AdditionalCostsPart2.
    Column mapping: CostName -> Original Cost Name; CostPrice/CostAmount -> Cost Price;
    CostCurrency -> Currency; PriceMechanism -> Rate by; ApplyTo -> Apply Over;
    CostCode -> Additional info(Cost Code); Validity Date -> Valid From; Carrier -> Carrier.
    Cost Type is filled by best-matching Original Cost Name against the 'Name' column from
    addition/Accessorial Costs file if cost_type_ref_path is provided. Apply if, Valid To left empty.
    """
    carrier = (metadata.get('carrier') or '').replace('\n', ' ')
    validity_date = metadata.get('validity_date', '')

    def item_to_row(item):
        cost_price = item.get('CostPrice') or item.get('CostAmount') or ''
        return {
            'Original Cost Name': item.get('CostName', ''),
            'Cost Type': '',
            'Cost Price': cost_price,
            'Currency': item.get('CostCurrency', ''),
            'Rate by': item.get('PriceMechanism', ''),
            'Apply Over': item.get('ApplyTo', ''),
            'Apply if': '',
            'Additional info(Cost Code)': item.get('CostCode', ''),
            'Valid From': validity_date,
            'Valid To': '',
            'Carrier': carrier,
        }

    rows = []
    for item in additional_costs_1 or []:
        rows.append(item_to_row(item))
    for item in additional_costs_2 or []:
        rows.append(item_to_row(item))

    if cost_type_ref_path is None:
        base = Path(__file__).resolve().parent
        for name in ('Accessorial Costs.xlsx', 'Accessorial Costs.csv'):
            p = base / 'addition' / name
            if p.exists():
                cost_type_ref_path = p
                break

    if cost_type_ref_path:
        name_list = _load_accessorial_cost_type_names(cost_type_ref_path)
        if name_list:
            for row in rows:
                original = row.get('Original Cost Name', '')
                row['Cost Type'] = _best_match_cost_type(original, name_list)

    return rows


def write_accessorial_sheet(workbook, sheet_name, rows):
    """Write the Accessorial Costs sheet with fixed column order."""
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    ws = workbook.create_sheet(sheet_name)
    columns = ACCESSORIAL_COSTS_COLUMNS

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


def save_to_excel(data, output_path):
    """Save all data to multi-tab Excel file"""
    print(f"[*] Creating Excel file: {output_path}")
    
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed!")
        print("        To install: pip install openpyxl")
        raise
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        metadata = data.get('metadata', {})
        
        # Tab 1: Metadata
        create_metadata_sheet(wb, metadata)
        
        # Tab 2: MainCosts (Matrix view)
        main_costs_data = data.get('MainCosts', [])
        zoning_matrix = data.get('ZoningMatrix', [])
        if main_costs_data:
            matrix_rows, category_specs = build_matrix_main_costs(main_costs_data, metadata)
            if zoning_matrix:
                matrix_rows = expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix)
            write_matrix_sheet(wb, "MainCosts", matrix_rows, category_specs, metadata)
        
        # Tab 3: AddedRates (pivoted)
        added_rates = data.get('AddedRates', [])
        if added_rates:
            added_rates_rows = pivot_added_rates(added_rates, metadata)
            write_sheet(wb, "AddedRates", added_rates_rows, metadata)
        
        # Tab 4: AdditionalCostsPart1
        additional_costs_1 = data.get('AdditionalCostsPart1', [])
        if additional_costs_1:
            additional_costs_1_rows = flatten_array_data(additional_costs_1, metadata, 'AdditionalCostsPart1')
            write_sheet(wb, "AdditionalCostsPart1", additional_costs_1_rows, metadata)
        
        # Tab 5: CountryZoning
        country_zoning = data.get('CountryZoning', [])
        if country_zoning:
            country_zoning_rows = flatten_array_data(country_zoning, metadata, 'CountryZoning')
            write_sheet(wb, "CountryZoning", country_zoning_rows, metadata)
        
        # Tab 6: AdditionalZoning
        additional_zoning = data.get('AdditionalZoning', [])
        if additional_zoning:
            additional_zoning_rows = flatten_array_data(additional_zoning, metadata, 'AdditionalZoning')
            write_sheet(wb, "AdditionalZoning", additional_zoning_rows, metadata)
        
        # Tab 7: ZoningMatrix
        zoning_matrix = data.get('ZoningMatrix', [])
        if zoning_matrix:
            zoning_matrix_rows = flatten_array_data(zoning_matrix, metadata, 'ZoningMatrix')
            write_sheet(wb, "ZoningMatrix", zoning_matrix_rows, metadata)
        
        # Tab 8: AdditionalCostsPart2
        additional_costs_2 = data.get('AdditionalCostsPart2', [])
        if additional_costs_2:
            additional_costs_2_rows = flatten_array_data(additional_costs_2, metadata, 'AdditionalCostsPart2')
            write_sheet(wb, "AdditionalCostsPart2", additional_costs_2_rows, metadata)
        
        # Tab 9: Accessorial Costs (combined from AdditionalCostsPart1 and AdditionalCostsPart2)
        accessorial_rows = build_accessorial_costs_rows(
            data.get('AdditionalCostsPart1', []),
            data.get('AdditionalCostsPart2', []),
            metadata,
        )
        if accessorial_rows:
            write_accessorial_sheet(wb, "Accessorial Costs", accessorial_rows)
        
        # Save workbook
        wb.save(output_path)
        
        file_size = os.path.getsize(output_path)
        file_size_kb = file_size / 1024
        
        print(f"[OK] Excel file saved successfully")
        print(f"  - Tabs: {len(wb.sheetnames)}")
        print(f"  - File size: {file_size_kb:.2f} KB")
        
    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")
        raise


def main():
    """Main execution function"""
    print("=" * 60)
    print("DHL RATE CARD EXCEL GENERATOR")
    print("=" * 60)
    print()
    
    # Define paths
    input_file = 'processing/extracted_data.json'
    output_dir = 'output'
    output_file = os.path.join(output_dir, 'DHL_Rate_Cards.xlsx')
    
    try:
        # Create output directory if it doesn't exist
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        print(f"[OK] Output directory ready: {output_dir}")
        print()
        
        # Load data
        print("Step 1: Loading extracted data...")
        data = load_extracted_data(input_file)
        print()
        
        # Save Excel
        print("Step 2: Creating multi-tab Excel file...")
        save_to_excel(data, output_file)
        print()
        
        # Success summary
        print("=" * 60)
        print("[SUCCESS] EXCEL GENERATION COMPLETE")
        print("=" * 60)
        print(f"Output file: {output_file}")
        print()
        print("Tabs created:")
        print("  1. Metadata (Carrier, Validity info)")
        
        stats = data.get('statistics', {})
        if stats.get('MainCosts_sections', 0) > 0:
            print(f"  2. MainCosts ({stats.get('MainCosts_rows', 0)} pricing rows)")
        if stats.get('AddedRates_rows', 0) > 0:
            print(f"  3. AddedRates ({stats.get('AddedRates_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart1_rows', 0) > 0:
            print(f"  4. AdditionalCostsPart1 ({stats.get('AdditionalCostsPart1_rows', 0)} rows)")
        if stats.get('CountryZoning_rows', 0) > 0:
            print(f"  5. CountryZoning ({stats.get('CountryZoning_rows', 0)} rows)")
        if stats.get('AdditionalZoning_rows', 0) > 0:
            print(f"  6. AdditionalZoning ({stats.get('AdditionalZoning_rows', 0)} rows)")
        if stats.get('ZoningMatrix_rows', 0) > 0:
            print(f"  7. ZoningMatrix ({stats.get('ZoningMatrix_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart2_rows', 0) > 0:
            print(f"  8. AdditionalCostsPart2 ({stats.get('AdditionalCostsPart2_rows', 0)} rows)")
        acc_count = len(data.get('AdditionalCostsPart1', [])) + len(data.get('AdditionalCostsPart2', []))
        if acc_count > 0:
            print(f"  9. Accessorial Costs ({acc_count} rows)")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print("[FAILED] EXCEL GENERATION FAILED")
        print("=" * 60)
        print(f"Error: {e}")
        print()
        raise


if __name__ == "__main__":
    main()



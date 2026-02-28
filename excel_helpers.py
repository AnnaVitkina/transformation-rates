"""
Excel sheet writers for the DHL rate-card workbook.

This module contains the three functions that actually write data into Excel sheets,
plus the ACCESSORIAL_COSTS_COLUMNS constant that defines the fixed column layout
for the Accessorial Costs tab.

Functions:
  write_matrix_sheet      – writes the MainCosts tab (special 3-row header)
  write_sheet             – writes any standard flat-table tab
  write_accessorial_sheet – writes the Accessorial Costs tab (fixed column order)

These functions are called by transformation_to_excel.py (the main orchestrator).
They do not transform data themselves; they only handle the Excel formatting and writing.
"""

import re
from pathlib import Path


# This list defines the exact columns and their order for the Accessorial Costs sheet.
# It is defined here as a constant so both the row-builder (in accessorial_costs.py)
# and the sheet-writer below use the same column order without having to pass it around.
ACCESSORIAL_COSTS_COLUMNS = [
    'Original Cost Name',          # the cost name as it appears in the rate card PDF
    'Cost Type',                   # standardised type name (filled by fuzzy matching)
    'Cost Price',                  # the numeric price value
    'Currency',                    # e.g. EUR, USD
    'Rate by',                     # how the price is applied (e.g. per shipment, per kg)
    'Apply Over',                  # what the cost applies to (e.g. base freight)
    'Apply if',                    # condition under which the cost applies (left blank)
    'Additional info(Cost Code)',  # internal cost code from the rate card
    'Valid From',                  # start date of validity (taken from the rate card metadata)
    'Valid To',                    # end date of validity (not available; left blank)
    'Carrier',                     # carrier name
]


def write_matrix_sheet(workbook, sheet_name, matrix_rows, category_specs, metadata):
    """
    Write the MainCosts tab to Excel with a special three-row header structure.

    WHY THREE HEADER ROWS?
    The MainCosts tab groups prices by cost category (e.g. "Documents", "Parcels").
    Each category has multiple weight columns (0.5 KG, 1 KG, 2 KG …).
    To make this readable, the header spans three rows:

      Row 1: Lane # | Origin | Destination | Service | Matrix zone | <-- Documents --> | <-- Parcels -->
      Row 2:        |        |             |         |             | Weight measure-KG |                |
      Row 3:        |        |             |         |             | 0.5 | 1 | 2 | 5   | 0.5 | 1 | 2 |
      Row 4+: actual data

    The category name in Row 1 is merged across all its weight columns.
    Data rows start at row 4.
    """
    if not matrix_rows:
        print(f"[WARN] No matrix data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} (Matrix) tab with {len(matrix_rows)} lanes...")
    ws = workbook.create_sheet(sheet_name)

    # Define the blue header style used for all three header rows
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # These five columns always appear first (left side of the sheet)
    fixed_cols = ['Lane #', 'Origin', 'Destination', 'Service', 'Matrix zone']
    num_fixed = len(fixed_cols)
    col = 1   # tracks the current column position as we build the header

    # --- Write the five fixed column names in Row 1 ---
    for c, name in enumerate(fixed_cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    col = num_fixed + 1   # move the column pointer past the fixed columns

    # --- Build the cost category column groups (Rows 1, 2, and 3) ---
    # For each cost category (e.g. "Documents"), we create a block of columns:
    #   - 1 extra column at the start of the block (for the "Weight measure - KG" label)
    #   - Then one column per weight breakpoint (0.5, 1, 2 …)
    category_start_cols = []   # stores position info for each category group

    for cost_cat_name, weight_unit, weights in category_specs:
        start_col = col   # remember where this category group starts

        # Row 2, first column of this group: write "Rate by: Weight measure - KG"
        # The "Rate by:" prefix is a fixed label; the rest is the extracted weight unit comment.
        _base_label = f"Weight measure - {weight_unit}" if weight_unit else "Weight measure"
        weight_measure_label = f"Rate by: {_base_label}"
        ws.cell(row=2, column=col, value=weight_measure_label)
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=2, column=col).font = header_font
        ws.cell(row=2, column=col).alignment = header_alignment
        col += 1

        # Row 2, remaining columns in this group: empty but styled (just the blue background)
        for _ in weights:
            ws.cell(row=2, column=col, value='')
            ws.cell(row=2, column=col).fill = header_fill
            col += 1

        # Row 3, first column of this group: empty (aligns with the "Weight measure" label above)
        ws.cell(row=3, column=start_col, value='')
        ws.cell(row=3, column=start_col).fill = header_fill
        col = start_col + 1

        # Row 3, remaining columns: write each weight breakpoint value (e.g. "<= 0.5", "<= 1", "<= 2")
        for w in weights:
            ws.cell(row=3, column=col, value=f"<= {w}")
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font
            ws.cell(row=3, column=col).alignment = header_alignment
            col += 1

        end_col = col - 1   # last column of this category group

        # Save the position info so we can write data rows correctly later
        category_start_cols.append((start_col, end_col, cost_cat_name, weight_unit, weights))

        # Row 1: merge all columns in this group into one cell and write the category name
        # e.g. "Documents" spans columns 6 to 10
        if start_col <= end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=cost_cat_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    total_cols = col - 1   # total number of columns in the sheet

    # Rows 2, 3 and 4 under the five fixed columns: empty cells with the blue header fill
    # (so the header band looks continuous across the full width)
    for r in (2, 3, 4):
        for c in range(1, num_fixed + 1):
            ws.cell(row=r, column=c, value='')
            ws.cell(row=r, column=c).fill = header_fill

    # Row 4 for each cost category group:
    #   - spacer column (first col of the group, below "Rate by: Weight measure - KG"): "Currency"
    #   - each weight column (below the weight breakpoint values): "Flat"
    for start_col, end_col, cost_cat_name, weight_unit, weights in category_start_cols:
        # Spacer column → "Currency" (sits directly below "Rate by: Weight measure - KG")
        cell = ws.cell(row=4, column=start_col, value='Currency')
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        # Weight columns → "Flat" (sits directly below each weight breakpoint value)
        for w_idx, _ in enumerate(weights):
            c = start_col + 1 + w_idx
            cell = ws.cell(row=4, column=c, value='Flat')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    # --- Write the data rows starting at row 5 (shifted down by one for the new Currency row) ---
    for row_idx, row_data in enumerate(matrix_rows, 5):
        col = 1

        # Write the five fixed columns (Lane #, Origin, Destination, Service, Matrix zone)
        for fc in fixed_cols:
            val = row_data.get(fc, '')
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fc == 'Lane #':
                cell.alignment = Alignment(horizontal="center")   # numbers look better centred
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            col += 1

        # Write the price columns for each cost category group.
        # The first column of each group is a spacer (empty).
        # The remaining columns each hold the price for one weight breakpoint.
        # The price is looked up from the row dict using the composite key (category, weight).
        for start_col, end_col, cost_cat_name, weight_unit, weights in category_start_cols:
            # First column of the group: spacer (empty)
            cell = ws.cell(row=row_idx, column=start_col, value='')
            cell.alignment = Alignment(horizontal="center")
            col = start_col + 1

            # One column per weight: look up the price and write it
            for w in weights:
                key = (cost_cat_name, w)   # e.g. ("Documents", "0.5")
                val = row_data.get(key, '')   # e.g. 10.50, or '' if no price for this weight
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                col += 1

    # --- Auto-size column widths ---
    # Sample the content of up to 53 rows (3 header rows + first 50 data rows)
    # to estimate a good column width.  Cap at 50 characters to avoid very wide columns.
    last_data_row = len(matrix_rows) + 3
    for c in range(1, total_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10   # minimum width
        for r in range(1, min(last_data_row + 1, 54)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze the first four rows so the header stays visible when scrolling down
    ws.freeze_panes = "A5"
    # Add a filter dropdown to row 4 (the Currency/Flat row) so users can filter
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{last_data_row}"
    print(f"[OK] {sheet_name} (Matrix) tab created with {total_cols} columns")


def write_sheet(workbook, sheet_name, rows, metadata):
    """
    Write a standard flat-table Excel sheet (used for AddedRates, CountryZoning,
    AdditionalZoning, ZoningMatrix, AdditionalCostsPart1, AdditionalCostsPart2).

    This is the generic writer used for all tabs except MainCosts (which has its own
    special three-row header).  It produces a simple one-row header + data rows layout.

    COLUMN ORDERING:
    Columns are arranged in three groups, in this order:
      1. Priority columns  – always appear first in a fixed human-friendly sequence
                             (Client, Carrier, Validity Date, Country, Country Code, …)
      2. Weight columns    – columns whose name contains "KG", starts with "<=", or contains "-"
                             sorted numerically (0.5 KG before 1 KG before 2 KG)
      3. Zone columns      – "Zone 1", "Zone 2" … sorted numerically
         Other columns     – everything else, sorted alphabetically
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(sheet_name)

    # Collect every column name that appears in any row (some rows may have extra fields)
    all_columns = set()
    for row in rows:
        all_columns.update(row.keys())

    # -----------------------------------------------------------------------
    # Step 1: Place the priority columns first.
    # These are the most important / most commonly used columns and should
    # always appear on the left side of the sheet.
    # -----------------------------------------------------------------------
    priority_cols = [
        'Client', 'Carrier', 'Validity Date',   # identity columns (always first)
        'Section', 'Service Type', 'Cost Category', 'Weight Unit', 'Zone',
        'Page Stopper', 'Table Name', 'Weight From', 'Weight To',
        'RateName', 'Country', 'Country Code', 'WeightFrom', 'WeightTo'
    ]

    columns = []
    for col in priority_cols:
        if col in all_columns:
            columns.append(col)
            all_columns.discard(col)   # remove from the remaining set so it doesn't appear twice

    # -----------------------------------------------------------------------
    # Step 2: From the remaining columns, separate weight columns from everything else.
    # Weight columns are identified by their name pattern:
    #   - Contains "KG"    e.g. "0.5 KG", "1 KG"
    #   - Starts with "<=" e.g. "<=0.5"
    #   - Contains "-"     e.g. "0-0.5"
    # -----------------------------------------------------------------------
    weight_cols = []
    other_cols = []

    for col in all_columns:
        if 'KG' in col or col.startswith('<=') or '-' in col:
            weight_cols.append(col)
        else:
            other_cols.append(col)

    # Sort weight columns numerically by the leading number
    # e.g. "0.5 KG", "1 KG", "2 KG" (not "0.5 KG", "2 KG", "1 KG")
    try:
        weight_cols_sorted = sorted(weight_cols, key=lambda x: float(x.split()[0]))
    except Exception:
        weight_cols_sorted = sorted(weight_cols)   # fallback: alphabetical

    # Sort "Zone N" columns numerically (Zone 1, Zone 2, Zone 10 …)
    # and sort all other columns alphabetically after them.
    def _other_col_sort_key(c):
        m = re.match(r'^Zone\s+(\d+)$', c, re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))   # group 0: sort by zone number
        return (1, c)                      # group 1: sort alphabetically

    columns.extend(weight_cols_sorted)
    columns.extend(sorted(other_cols, key=_other_col_sort_key))

    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write the header row (row 1) with the column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows starting at row 2.
    # For each row, look up the value for each column and write it to the correct cell.
    # If a row doesn't have a value for a column, write an empty string.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Short numeric/code values are centred; longer text values wrap inside the cell
            if column in ['Weight', 'Weight Unit', 'Section', 'Zone', 'Currency', 'Rate'] or 'KG' in column:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size column widths by looking at the content of the first 50 data rows.
    # The width is capped between 10 and 50 characters to avoid extremes.
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length as the minimum
        for row_idx in range(2, min(len(rows) + 2, 52)):   # sample up to 50 data rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze the header row so column names stay visible when scrolling down
    ws.freeze_panes = "A2"
    # Add filter dropdowns to every column so users can filter/sort the data
    ws.auto_filter.ref = ws.dimensions

    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


def write_accessorial_sheet(workbook, sheet_name, rows):
    """
    Write the Accessorial Costs tab to Excel.

    This is a simplified version of write_sheet() that uses the fixed column order
    defined in ACCESSORIAL_COSTS_COLUMNS instead of dynamically determining columns.
    The column order is fixed because the Accessorial Costs tab has a specific agreed layout.
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    ws = workbook.create_sheet(sheet_name)
    columns = ACCESSORIAL_COSTS_COLUMNS   # use the fixed column list defined at the top of this file

    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write the header row (row 1) with the fixed column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows starting at row 2.
    # All cells use wrap_text so long cost names are readable without widening the column too much.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')   # empty string if this row has no value for this column
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns by sampling up to 100 data rows (more than write_sheet's 50,
    # because cost names can be long and we want to capture outliers)
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

    # Freeze the header row and add filter dropdowns
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")

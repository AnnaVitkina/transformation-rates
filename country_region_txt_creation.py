"""
Read DHL_Rate_Cards.xlsx, tab CountryZoning; write a TXT file that lists
each RateName followed by all countries with that rate name (comma-separated).
Output is saved to the output folder.
"""

from pathlib import Path
from collections import defaultdict


def create_country_region_txt(
    excel_path: str = "output/DHL_Rate_Cards.xlsx",
    sheet_name: str = "CountryZoning",
    output_path: str | None = None,
) -> str:
    """
    Load CountryZoning from Excel, group countries by RateName, write TXT.
    Returns the path of the created file.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    excel_path = Path(excel_path)
    print(f"[*] TXT Debug: excel_path={excel_path}")
    print(f"[*] TXT Debug: sheet_name={sheet_name}")
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if output_path is None:
        output_dir = excel_path.parent  # output folder
        output_path = output_dir / "CountryZoning_by_RateName.txt"
    else:
        output_path = Path(output_path)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("", encoding="utf-8")
        print(f"[WARN] Sheet '{sheet_name}' not found in {excel_path} (no CountryZoning data in this rate card). Wrote empty TXT: {output_path}")
        return str(output_path)
    print(f"[*] TXT Debug: workbook sheets={wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print(f"[*] TXT Debug: total rows read (including header)={len(rows)}")

    if not rows:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("", encoding="utf-8")
        print("[WARN] TXT Debug: sheet is empty, wrote empty txt")
        return str(output_path)

    # First row = headers; find column indices for RateName and Country
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    rate_name_col = None
    country_col = None
    for i, h in enumerate(headers):
        if h == "RateName":
            rate_name_col = i
        if h == "Country Code":
            country_col = i
    print(f"[*] TXT Debug: headers={headers}")
    print(f"[*] TXT Debug: RateName col index={rate_name_col}, Country Code col index={country_col}")
    if rate_name_col is None:
        raise ValueError("Column 'RateName' not found in CountryZoning")
    if country_col is None:
        raise ValueError("Column 'Country Code' not found in CountryZoning")

    # Group countries by RateName. Forward-fill RateName (empty cells = same as previous row).
    by_rate_name = defaultdict(list)
    current_rate = ""
    processed_rows = 0
    skipped_empty_country = 0
    for row in rows[1:]:
        rn = row[rate_name_col] if rate_name_col < len(row) else None
        country = row[country_col] if country_col < len(row) else None
        if rn is not None and str(rn).strip():
            current_rate = str(rn).strip()
        if country is None or (isinstance(country, str) and not str(country).strip()):
            skipped_empty_country += 1
            continue
        country_str = str(country).strip()
        if country_str:
            by_rate_name[current_rate].append(country_str)
            processed_rows += 1
    print(f"[*] TXT Debug: processed country rows={processed_rows}")
    print(f"[*] TXT Debug: skipped rows with empty Country Code={skipped_empty_country}")
    print(f"[*] TXT Debug: distinct RateName groups={len(by_rate_name)}")

    # Write TXT: "RateName - country1, country2, ..."
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    for rate_name in sorted(by_rate_name.keys(), key=lambda x: (x == "", x)):
        countries = by_rate_name[rate_name]
        line = f"{rate_name}  {', '.join(countries)}"
        lines.append(line)
    print(f"[*] TXT Debug: output lines={len(lines)}")
    if lines:
        print(f"[*] TXT Debug: first line preview={lines[0][:200]}")
    else:
        print("[WARN] TXT Debug: no lines generated, output will be empty")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[OK] TXT Debug: wrote file {output_path}")
    return str(output_path)


def main():
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / "output" / "DHL_Rate_Cards.xlsx"
    output_path = script_dir / "output" / "CountryZoning_by_RateName.txt"

    print("Creating CountryZoning TXT from DHL_Rate_Cards.xlsx...")
    out = create_country_region_txt(
        excel_path=str(excel_path),
        output_path=str(output_path),
    )
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()

